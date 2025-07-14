"""
El País Opinion Articles Scraper – BrowserStack Version
Author: Ridhi Moda
"""

import os
import sys
import time
import logging
import requests
from collections import Counter
from datetime import datetime
from urllib.parse import urljoin

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.safari.options import Options as SafariOptions

from googletrans import Translator

# ← Load your .env so the credentials are in os.environ
from dotenv import load_dotenv
load_dotenv()

# ← BrowserStack Local tunnel helper
from browserstack.local import Local

from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock


class ElPaisLocalScraper:
    """
    Local scraper for El País Opinion section
    """
    
    def __init__(self, base_dir=None):
        """Initialize scraper with directory setup"""
        self.translator = Translator()
        self.base_url = "https://elpais.com"
        self.opinion_url = "https://elpais.com/opinion/"
        
        # Setup directories
        self.base_dir = base_dir or os.getcwd()
        self.images_dir = os.path.join(self.base_dir, "article_images")
        os.makedirs(self.images_dir, exist_ok=True)
        
        # Setup logging
        log_filename = os.path.join(
            self.base_dir,
            f'execution_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        )
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Local scraper initialized. Base directory: {self.base_dir}")
        
        self.articles_data = []
        self.seen_titles = set()  # Track unique titles
    
    def create_driver(self):
        """Create a local Chrome WebDriver instance"""
        chrome_options = Options()
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        driver = webdriver.Chrome(options=chrome_options)
        self.logger.info("Chrome driver created successfully")
        return driver
    
    def download_image(self, image_url, article_index):
        """Download article cover image"""
        try:
            self.logger.info(f"Downloading image for article {article_index}: {image_url}")
            resp = requests.get(
                image_url, timeout=10,
                headers={'User-Agent': 'Mozilla/5.0'}
            )
            if resp.status_code == 200:
                filename = f"article_{article_index}_cover.jpg"
                filepath = os.path.join(self.images_dir, filename)
                with open(filepath, 'wb') as f:
                    f.write(resp.content)
                self.logger.info(f"Successfully downloaded image: {filepath}")
                return filepath
            else:
                self.logger.warning(f"Failed to download image: {resp.status_code}")
        except Exception as e:
            self.logger.error(f"Error downloading image: {e}")
        return None
    
    def get_article_summary(self, article_element):
        """Get article summary from the main page"""
        selectors = ["p.c_d","p.c-d__sumario","p",".c_d",".description"]
        for sel in selectors:
            try:
                elem = article_element.find_element(By.CSS_SELECTOR, sel)
                txt  = elem.text.strip()
                if txt and len(txt) > 20:
                    return txt
            except:
                continue
        return "Summary not available"
    
    def scrape_articles(self, driver):
        """Main scraping logic - Opinion section only"""
        try:
            self.logger.info("Starting scraping process...")
            driver.get(self.opinion_url)
            time.sleep(20)
            
            # Handle cookie consent
            try:
                btn = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.ID,"didomi-notice-agree-button"))
                )
                btn.click(); time.sleep(12)
                self.logger.info("Cookie consent handled")
            except:
                self.logger.info("No cookie banner")
            
            # Screenshot landing page
            shot = os.path.join(self.base_dir,"opinion_page_screenshot.png")
            driver.save_screenshot(shot)
            self.logger.info(f"Screenshot saved: {shot}")
            
            # Find up to 5 articles
            selectors = ["article.c_c","article.c","div.c_c article","section article"]
            articles = []
            for sel in selectors:
                found = driver.find_elements(By.CSS_SELECTOR, sel)
                if found:
                    articles = found
                    self.logger.info(f"Found {len(articles)} articles using {sel}")
                    break
            if not articles:
                articles = driver.find_elements(By.TAG_NAME, "article")
                self.logger.info(f"Found {len(articles)} articles by <article> tag")
            
            count = 0
            for art in articles:
                if count >= 5:
                    break
                try:
                    text = art.text.strip()
                    if len(text) < 50:
                        continue
                    data = {
                        'article_number': count+1,
                        'spanish_title': '',
                        'english_title': '',
                        'spanish_content': '',
                        'english_content': '',
                        'image_url': '',
                        'image_path': '',
                        'article_url': ''
                    }
                    # Title extraction
                    t_selectors = [
                        "h2.c_t","h2 a","h3.c_t","h3 a",
                        "header h2","header h3",".c_t"
                    ]
                    for ts in t_selectors:
                        try:
                            te = art.find_element(By.CSS_SELECTOR, ts)
                            txt = te.text.strip()
                            if (txt 
                                and txt not in [
                                    'EDITORIAL','TRIBUNA','COLUMNA',
                                    'CARTAS AL DIRECTOR','EXPOSICIÓN'
                                ] 
                                and txt not in self.seen_titles
                            ):
                                data['spanish_title'] = txt
                                self.seen_titles.add(txt)
                                if te.tag_name == 'a':
                                    data['article_url'] = te.get_attribute("href")
                                else:
                                    try:
                                        link = art.find_element(By.CSS_SELECTOR,"a")
                                        data['article_url'] = link.get_attribute("href")
                                    except:
                                        pass
                                break
                        except:
                            continue
                    if not data['spanish_title']:
                        continue
                    
                    data['spanish_content'] = self.get_article_summary(art)
                    # Pick up image URL
                    try:
                        ie = art.find_element(By.CSS_SELECTOR,"img")
                        url = ie.get_attribute("src") or ie.get_attribute("data-src")
                        if url and not url.startswith("http"):
                            url = urljoin(self.base_url, url)
                        data['image_url'] = url
                    except:
                        self.logger.info("No image for this article")
                    
                    self.logger.info(f"Artículo {count+1}: {data['spanish_title']}")
                    self.articles_data.append(data)
                    count += 1
                except Exception as e:
                    self.logger.error(f"Error processing article: {e}")
            self.logger.info(f"Total scraped: {len(self.articles_data)}")
            return True
        except Exception as e:
            self.logger.error(f"Scraping error: {e}")
            return False
    
    def download_all_images(self):
        """Download all article images"""
        self.logger.info("Downloading images...")
        for art in self.articles_data:
            if art['image_url']:
                path = self.download_image(
                    art['image_url'], art['article_number']
                )
                if path:
                    art['image_path'] = path
    
    def translate_content(self):
        """Translate Spanish → English"""
        self.logger.info("Starting translation...")
        for art in self.articles_data:
            try:
                if art['spanish_title']:
                    t = self.translator.translate(
                        art['spanish_title'], src='es', dest='en'
                    )
                    art['english_title'] = t.text
                    self.logger.info(
                        f"Title: {art['spanish_title']} → {t.text}"
                    )
                if art['spanish_content']:
                    txt = art['spanish_content']
                    if len(txt) > 1000:
                        txt = txt[:1000] + "..."
                    t = self.translator.translate(
                        txt, src='es', dest='en'
                    )
                    art['english_content'] = t.text
            except Exception as e:
                self.logger.error(f"Translation error: {e}")
                art['english_title']   = "Translation failed"
                art['english_content'] = "Translation failed"
                time.sleep(1)

    def analyze_word_frequency(self):
        """Analyze word frequency from translated titles"""
        self.logger.info("\n=== WORD FREQUENCY ANALYSIS ===")
        
        all_words = []
        for article in self.articles_data:
            if article.get('english_title') and article['english_title'] != "Translation failed":
                words = article['english_title'].lower().split()
                # Remove common stop words
               #  stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'is', 'are', 'was', 'were', 'it', 'that', 'this', 'as'}
              # words = [word.strip('.,!?;:"\'-()[]{}') for word in words]
                all_words.extend(words)
        
        word_counts = Counter(all_words)
        repeated_words = {word: count for word, count in word_counts.items() if count > 2}
        
        # Log results
        if repeated_words:
            self.logger.info("Words repeated more than twice:")
          #  for word, count in sorted(repeated_words.items(), key=lambda x: x[1], reverse=True):
            for word, count in repeated_words.items():
                self.logger.info(f"- '{word}': {count} times")
        else:
            self.logger.info("No words repeated more than twice")
        
        return word_counts, repeated_words
    
    # def analyze_word_frequency(self):
    #     """Analyze title word frequency and report words repeated more than twice."""
    #     self.logger.info("Word frequency analysis...")
        
    #     # Collect all words from translated titles
    #     all_words = []
    #     #  ##stop_words = {
    #     #     'the','a','an','and','or','but','in','on','at','to',
    #     #     'for','of','with','by','is','are','was','were','it',
    #     #     'that','this','as'
    #     # }
        
    #     for article in self.articles_data:
    #         title = article.get('english_title', '').lower()
    #         if title and title != "Translation failed":
    #             words = [
    #                 w.strip('.,!?;:"\'-()[]{}')
    #                 for w in title.split()
    #                 if w not in stop_words and len(w) > 2
    #             ]
    #             all_words.extend(words)
        
    #     # Count frequencies
    #     counts = Counter(all_words)
        
    #     # Filter words with frequency > 2
    #     repeats = {word: freq for word, freq in counts.items() if freq > 2}
        
    #     # Report to console and log file
    #     if repeats:
    #         print("\n🔤 Words repeated more than twice:")
    #         for word, freq in repeats.items():
    #             print(f"• {word} — {freq} times")
            
    #         self.logger.info("Words repeated more than twice and their counts:")
    #         for word, freq in repeats.items():
    #             self.logger.info(f"{word}: {freq}")
    #     else:
    #         print("\n🔤 No words were repeated more than twice.")
    #         self.logger.info("No words repeated more than twice.")
        
    #     return counts, repeats
    
    def create_excel_report(self, word_analysis):
        """Generate Bain-style Excel"""
        self.logger.info("Creating Excel report...")
        path = os.path.join(
            self.base_dir,
            f'ElPais_Analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        wb = Workbook()
        # — Styles & sheets exactly as before —
        # Sheet1: Article Details
        ws1 = wb.active
        ws1.title = "Article Details"
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(
            start_color='1F4E79', end_color='1F4E79', fill_type='solid'
        )
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cols = [
            'Article Number','Spanish Title','English Title',
            'Spanish Content','English Content'
        ]
        for idx, h in enumerate(cols,1):
            c = ws1.cell(row=1, column=idx, value=h)
            c.font, c.fill, c.alignment, c.border = (
                header_font, header_fill, header_align, thin
            )
        for r, art in enumerate(self.articles_data, start=2):
            ws1.cell(row=r, column=1, value=art['article_number']).border = thin
            ws1.cell(row=r, column=2, value=art['spanish_title']).border = thin
            ws1.cell(row=r, column=3, value=art['english_title']).border = thin
            ws1.cell(row=r, column=4, value=art['spanish_content']).border = thin
            ws1.cell(row=r, column=5, value=art['english_content']).border = thin
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 45
        ws1.column_dimensions['C'].width = 45
        ws1.column_dimensions['D'].width = 70
        ws1.column_dimensions['E'].width = 70
        for r in range(2, len(self.articles_data)+2):
            ws1.row_dimensions[r].height = 80
        
        # Sheet2: Word Frequency Analysis
        ws2 = wb.create_sheet("Word Frequency Analysis")
        ws2.merge_cells('A1:C1')
        tcell = ws2['A1']
        tcell.value = "Word Frequency Analysis – Repeated Words (>2)"
        tcell.font = Font(name='Arial', size=14, bold=True)
        tcell.alignment = Alignment(horizontal='center')
        headers = ['Rank','Word','Frequency']
        for idx, h in enumerate(headers,1):
            c = ws2.cell(row=3, column=idx, value=h)
            c.font, c.fill, c.alignment, c.border = (
                header_font, header_fill, header_align, thin
            )
        counts, repeats = word_analysis
        items = sorted(repeats.items(), key=lambda x: x[1], reverse=True)
        for i,(w,cnt) in enumerate(items, start=1):
            row = i+3
            ws2.cell(row=row, column=1, value=i).border = thin
            ws2.cell(row=row, column=2, value=w).border = thin
            ws2.cell(row=row, column=3, value=cnt).border = thin
        stats_row = len(items)+6 if items else 7
        ws2.cell(row=stats_row, column=1, value="Total Unique Words:").font = Font(bold=True)
        ws2.cell(row=stats_row, column=2, value=len(counts))
        ws2.cell(row=stats_row+1, column=1, value="Words Repeated >2:").font = Font(bold=True)
        ws2.cell(row=stats_row+1, column=2, value=len(repeats))
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 15
        
        wb.save(path)
        self.logger.info(f"Excel report saved: {path}")
        return path

# ------------------------------------------------------------------------------
# BrowserStack + Local integration 
# ------------------------------------------------------------------------------
BROWSERSTACK_USERNAME = os.getenv("BROWSERSTACK_USERNAME")
BROWSERSTACK_ACCESS_KEY = os.getenv("BROWSERSTACK_ACCESS_KEY")
if not BROWSERSTACK_USERNAME or not BROWSERSTACK_ACCESS_KEY:
    raise EnvironmentError(
        "Please set BROWSERSTACK_USERNAME & BROWSERSTACK_ACCESS_KEY in your .env"
    )

HUB_URL    = f"https://{BROWSERSTACK_USERNAME}:{BROWSERSTACK_ACCESS_KEY}@hub.browserstack.com/wd/hub"
BUILD_NAME = "ElPais_Opinion_Scraper"

# 5 parallel sessions
SESSIONS = [
    # Desktop
    {"browser": "Chrome",  "os": "Windows", "osVersion": "11"},
    {"browser": "Firefox", "os": "OS X",    "osVersion": "Ventura"},
    {"browser": "Edge",    "os": "Windows", "osVersion": "10"},

    # Mobile (real devices)
    {
      "browser":    "Safari",
      "deviceName": "iPhone 14",
      "realMobile": True,
      "osVersion":  "16"
    },
    {
      "browser":    "Chrome",
      "deviceName": "Samsung Galaxy S21",
      "realMobile": True,
      "osVersion":  "12"
    }
]
results_lock   = Lock()
shared_results = []

def start_local_tunnel():
    
    bs_local = Local()
    local_id = f"LocalTest_{int(time.time())}"
    bs_args = {
        "key":               BROWSERSTACK_ACCESS_KEY,
        "localIdentifier":   local_id,
        "forcelocal":        True
    }
    bs_local.start(**bs_args)
    print(f"✅ BrowserStack Local tunnel started (ID: {local_id})")
    return bs_local, local_id

from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.safari.options import Options as SafariOptions
import traceback
from selenium.common.exceptions import WebDriverException

def create_bs_driver(session_info, local_id):
    """
    Build a BrowserStack Remote driver with comprehensive error handling
    """
    session_name = None
    try:
        # Generate session name for logging
        platform = session_info.get("os") or session_info.get("deviceName") or "Unknown"
        session_name = f"{session_info['browser']}_{platform}_{session_info['osVersion']}"
        
        print(f"🚀 Creating BrowserStack driver for: {session_name}")
        
        # 1) Pick the right Options class
        browser = session_info["browser"].lower()
        if "chrome" in browser:
            opts = ChromeOptions()
        elif browser == "firefox":
            opts = FirefoxOptions()
        elif "edge" in browser:
            opts = EdgeOptions()
        elif "safari" in browser:
            opts = SafariOptions()
        else:
            print(f"⚠️  Unknown browser {browser}, defaulting to Chrome options")
            opts = ChromeOptions()

        # 2) Build bstack:options payload
        bstack_opts = {
            "buildName": BUILD_NAME,
            "sessionName": session_name,
            "local": "true",
            "localIdentifier": local_id,
            "debug": "true",
            "networkLogs": "true",
            "consoleLogs": "verbose"
        }

        # 3) Add browser version if specified
        if "browserVersion" in session_info:
            bstack_opts["browserVersion"] = session_info["browserVersion"]

        # 4) Desktop vs mobile configuration
        if "os" in session_info:
            # Desktop configuration
            bstack_opts["os"] = session_info["os"]
            bstack_opts["osVersion"] = session_info["osVersion"]
            print(f"🖥️  Desktop config: {session_info['os']} {session_info['osVersion']}")
        else:
            # Mobile configuration
            bstack_opts["deviceName"] = session_info["deviceName"]
            bstack_opts["realMobile"] = session_info.get("realMobile", True)
            bstack_opts["osVersion"] = session_info["osVersion"]
            print(f"📱 Mobile config: {session_info['deviceName']} {session_info['osVersion']}")

        # 5) Set capabilities
        try:
            opts.set_capability("bstack:options", bstack_opts)
            opts.set_capability("browserName", session_info["browser"])
            
            # Additional capabilities for stability
            if browser == "chrome":
                opts.add_argument("--no-sandbox")
                opts.add_argument("--disable-dev-shm-usage")
                opts.add_argument("--disable-gpu")
            
            print(f"🔧 Capabilities set for {session_name}")
            
        except Exception as e:
            print(f"❌ Error setting capabilities for {session_name}: {e}")
            raise

        # 6) Create remote driver
        try:
            print(f"🌐 Connecting to BrowserStack hub...")
            driver = webdriver.Remote(
                command_executor=HUB_URL,
                options=opts
            )
            
            print(f"✅ BrowserStack driver created successfully for: {session_name}")
            return driver
            
        except WebDriverException as e:
            print(f"❌ WebDriver error creating {session_name}: {e}")
            # Try to extract more specific error info
            if "Invalid capabilities" in str(e):
                print(f"🔍 Capabilities issue for {session_name}: {bstack_opts}")
            elif "Authentication" in str(e):
                print(f"🔍 Authentication issue - check credentials")
            elif "timeout" in str(e).lower():
                print(f"🔍 Connection timeout for {session_name}")
            raise
        
    except Exception as e:
        print(f"❌ Critical error creating driver for {session_name or 'unknown session'}: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        raise

# def create_bs_driver(session_info, local_id):
#     """
#     Build a BrowserStack Remote driver using an Options instance
#     so that Remote(options=opts) works in Selenium 4.15+.
#     """
#     # 1) Pick the right Options class
#     browser = session_info["browser"].lower()
#     if "chrome" in browser:
#         opts = ChromeOptions()
#     elif browser == "firefox":
#         opts = FirefoxOptions()
#     elif "edge" in browser:
#         opts = EdgeOptions()
#     elif "safari" in browser:
#         opts = SafariOptions()
#     else:
#         opts = ChromeOptions()  # fallback

#     # 2) Build your bstack:options payload
#     bstack_opts = {
#         "buildName":      BUILD_NAME,
#         "sessionName":    f"{session_info['browser']}_{session_info.get('os', session_info.get('deviceName'))}_{session_info['osVersion']}",
#         "local":          "true",
#         "localIdentifier": local_id
#     }

#     # 3) Desktop vs real-mobile keys
#     if "os" in session_info:
#         bstack_opts["os"]        = session_info["os"]
#         bstack_opts["osVersion"] = session_info["osVersion"]
#     else:
#         bstack_opts["deviceName"] = session_info["deviceName"]
#         bstack_opts["realMobile"] = session_info["realMobile"]
#         bstack_opts["osVersion"]  = session_info["osVersion"]

#     # 4) Attach to the Options object
#     opts.set_capability("bstack:options", bstack_opts)
#     opts.set_capability("browserName", session_info["browser"])

#     # 5) Launch the remote session
#     return webdriver.Remote(
#         command_executor=HUB_URL,
#         options=opts
#     )

def session_task(session_info, base_dir, local_id):
    """
    Run one BrowserStack session: scrape, download, translate,
    take screenshot, and aggregate results safely.
    """
    # Determine platform (use 'os' for desktop, 'deviceName' for mobile)
    platform = session_info.get("os") or session_info.get("deviceName") or "Unknown"
    session_name = f"{session_info['browser']}_{platform}_{session_info['osVersion']}"

    # Initialize your original scraper
    scraper = ElPaisLocalScraper(base_dir=base_dir)

    # Create a BrowserStack driver (uses create_bs_driver as defined)
    driver = create_bs_driver(session_info, local_id)

    try:
        # 1) Scrape articles
        success = scraper.scrape_articles(driver)
        if not success:
            scraper.logger.error(f"[{session_name}] Scrape failed, skipping")
            return

        # 2) Download images and translate
        scraper.download_all_images()
        scraper.translate_content()

        # 3) Capture landing-page screenshot
        shots_dir = os.path.join(base_dir, "browserstack_screenshots")
        os.makedirs(shots_dir, exist_ok=True)
        driver.get(scraper.opinion_url)
        time.sleep(10)
        shot_path = os.path.join(shots_dir, f"{session_name}.png")
        driver.save_screenshot(shot_path)
        scraper.logger.info(f"[{session_name}] Screenshot saved: {shot_path}")

        # 4) Aggregate results and safely rename images
        with results_lock:
            for art in scraper.articles_data:
                orig = art.get("image_path", "")
                if orig and os.path.exists(orig):
                    dirname, fname = os.path.split(orig)
                    new_fname = f"{session_name}__{fname}"
                    new_path  = os.path.join(dirname, new_fname)

                    if not os.path.exists(new_path):
                        try:
                            os.rename(orig, new_path)
                            art["image_path"] = new_path
                        except Exception as e:
                            scraper.logger.error(f"[{session_name}] Error renaming {orig} → {new_path}: {e}")
                    else:
                        scraper.logger.warning(f"[{session_name}] Skipping rename, target exists: {new_path}")
                else:
                    scraper.logger.warning(f"[{session_name}] Image not found, cannot rename: {orig}")

                # Tag with session name and add to shared list
                art["session"] = session_name
                shared_results.append(art)

    finally:
        driver.quit()
        scraper.logger.info(f"[{session_name}] Session closed")

def main():
    base_dir = sys.argv[1] if len(sys.argv)>1 else r"C:\Ridhi_Moda"
    os.makedirs(base_dir, exist_ok=True)

    # 1) Start the BrowserStack Local tunnel
    bs_local, local_id = start_local_tunnel()

    # 2) Run 5 sessions in parallel
    try:
        print(f"Running {len(SESSIONS)} BrowserStack sessions…")
        with ThreadPoolExecutor(max_workers=5) as exe:
            futs = [
                exe.submit(
                    session_task, sess, base_dir, local_id
                ) for sess in SESSIONS
            ]
            for f in as_completed(futs):
                f.result()
    finally:
        bs_local.stop()
        print("✅ BrowserStack Local tunnel stopped")

    # 3) Merge results & create your final Excel
    master = ElPaisLocalScraper(base_dir=base_dir)
    master.articles_data = shared_results
    word_analysis = master.analyze_word_frequency()
    excel_path    = master.create_excel_report(word_analysis)

    # 4) Final summary
    print("\n" + "="*60)
    print("✅ All sessions complete")
    print(f"📄 Excel report: {excel_path}")
    print(f"🖼️  Article images: {os.path.join(base_dir,'article_images')}")
    print(f"📸 Screenshots: {os.path.join(base_dir,'browserstack_screenshots')}")
    print("="*60)

if __name__ == "__main__":
    main()
